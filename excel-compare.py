import openpyxl
import tkinter as tk
from tkinter import filedialog

def compare_workbooks(file1, file2, output_file):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    
    discrepancies = []

    for sheet_name in wb1.sheetnames:
        if sheet_name in wb2.sheetnames:
            sheet1 = wb1[sheet_name]
            sheet2 = wb2[sheet_name]

            for row in range(1, sheet1.max_row + 1):
                for col in range(1, sheet1.max_column + 1):
                    cell1 = sheet1.cell(row=row, column=col)
                    cell2 = sheet2.cell(row=row, column=col)

                    if cell1.value != cell2.value:
                        discrepancies.append((sheet_name, row, col, cell1.value, cell2.value))
                        discrepancies.append((sheet_name, row, col, cell1.value, cell2.value))

    with open(output_file, 'w') as f:
        f.write("Sheet, Row, Column, Value in File1, Value in File2\n")
        for sheet, row, col, val1, val2 in discrepancies:
            f.write(f"{sheet}, {row}, {col}, {val1}, {val2}\n")

    print(f"Comparison report saved as {output_file}")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window

    print("Select the first Excel file to compare:")
    file1 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    print("Select the second Excel file to compare:")
    file2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file1 and file2:
        output_file = "comparison_report.csv"
        compare_workbooks(file1, file2, output_file)
    else:
        print("Files not selected. Exiting.")
